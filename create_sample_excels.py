import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_sample_excels():
    output_dirs = [
        r"c:\laragon\www\evaluaciones",
        r"c:\laragon\www\evaluaciones\evaluaciones-frontend\public\samples"
    ]
    
    for d in output_dirs:
        os.makedirs(d, exist_ok=True)

    headers = [
        'tipo', 'grupo', 'enunciado', 'opcion_a', 'opcion_b', 
        'opcion_c', 'opcion_d', 'opcion_e', 'respuesta_correcta', 
        'dificultad', 'peso', 'observaciones'
    ]

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill_purple = PatternFill(start_color='4527A0', end_color='4527A0', fill_type='solid')
    header_fill_cyan = PatternFill(start_color='006064', end_color='006064', fill_type='solid')
    header_fill_red = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_top_wrap = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D7DE'),
        right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'),
        bottom=Side(style='thin', color='D0D7DE')
    )

    # =========================================================================
    # 1. EXCEL VÁLIDO (60 PREGUNTAS: 15 FÁCILES, 30 MEDIAS, 15 DIFÍCILES)
    # =========================================================================
    wb_valido = openpyxl.Workbook()
    
    # Sheet 1: Instrucciones
    ws_inst = wb_valido.active
    ws_inst.title = "Instrucciones"
    ws_inst.column_dimensions['A'].width = 38
    ws_inst.column_dimensions['B'].width = 85
    
    ws_inst.append(["BANCO DE PREGUNTAS - GUÍA OFICIAL"])
    ws_inst.cell(row=1, column=1).font = Font(name='Calibri', size=15, bold=True, color='FFFFFF')
    ws_inst.cell(row=1, column=1).fill = header_fill_purple
    ws_inst.append([])
    ws_inst.append(["1. CÓDIGOS DE PREGUNTA", "FALSO_VERDADERO, PREGUNTA_CON_CLAVE, SELECCION_SIMPLE, RESPUESTA_COMPUESTA, PROBLEMA, SUBPROBLEMA, EMPAREJAMIENTO"])
    ws_inst.append(["2. CUOTAS OBLIGATORIAS 2DO PARCIAL", "15 Fáciles (1), 30 Medias (2), 15 Difíciles (3) = 60 preguntas evaluables"])
    ws_inst.append(["3. NOTAS DE FORMATO", "La columna observaciones muestra OK si el reactivo cumple la totalidad de reglas."])

    # Sheet 2: Banco (60 Filas)
    ws_banco = wb_valido.create_sheet(title="Banco")
    ws_banco.views.sheetView[0].showGridLines = True
    ws_banco.freeze_panes = 'A2'
    ws_banco.append(headers)
    for col_num in range(1, len(headers) + 1):
        c = ws_banco.cell(row=1, column=col_num)
        c.font = header_font
        c.fill = header_fill_purple
        c.alignment = align_center

    col_widths = [22, 14, 55, 25, 25, 25, 25, 25, 18, 12, 10, 25]
    for idx, width in enumerate(col_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(idx)
        ws_banco.column_dimensions[col_letter].width = width

    # 15 Fáciles (Dificultad 1)
    faciles_data = [
        ("FALSO_VERDADERO", "", "La fibra óptica monomodo presenta menor atenuación que la multimodo a distancias largas.", "Verdadero", "Falso", "", "", "", "A", "1", 5, "OK"),
        ("SELECCION_SIMPLE", "", "¿Cuál es la función principal de la capa de enlace de datos en el modelo OSI?", "Direccionamiento físico (MAC) y control de flujo", "Enrutamiento de paquetes", "Cifrado de datos", "Control de sesiones", "Compresión", "A", "1", 5, "OK"),
        ("FALSO_VERDADERO", "", "El protocolo UDP es no orientado a conexión y no garantiza la entrega de paquetes.", "Verdadero", "Falso", "", "", "", "A", "1", 5, "OK"),
        ("SELECCION_SIMPLE", "", "¿Qué longitud de onda se utiliza en la 3ra ventana de comunicaciones ópticas?", "850 nm", "1310 nm", "1550 nm", "1625 nm", "1490 nm", "C", "1", 5, "OK"),
        ("RESPUESTA_COMPUESTA", "", "I. HTTP opera sobre el puerto 80 TCP.\nII. DNS opera sobre el puerto 53 UDP/TCP.", "A. Si la primera es verdadera", "B. Si la segunda es verdadera", "C. Si ambas son verdaderas", "D. Si ninguna es verdadera", "", "C", "1", 5, "OK"),
        ("FALSO_VERDADERO", "", "La modulación AM varía la amplitud de la onda portadora manteniendo su frecuencia.", "Verdadero", "Falso", "", "", "", "A", "1", 5, "OK"),
        ("SELECCION_SIMPLE", "", "¿Cuál es la impedancia característica estándar del cable coaxial RG-6?", "50 Ohmios", "75 Ohmios", "100 Ohmios", "120 Ohmios", "300 Ohmios", "B", "1", 5, "OK"),
        ("PREGUNTA_CON_CLAVE", "", "Son medios no guiados de transmisión: 1. Ondas de radio, 2. Microondas, 3. Infrarrojos, 4. Cable STP.", "1, 2 y 3 son correctas", "1 y 3 son correctas", "2 y 4 son correctas", "Solo 4 es correcta", "Todas son correctas", "A", "1", 5, "OK"),
        ("SELECCION_SIMPLE", "", "¿Qué conector se utiliza habitualmente en cables UTP Categoría 6 para redes LAN?", "RJ-11", "RJ-45", "BNC", "SC/APC", "SMA", "B", "1", 5, "OK"),
        ("FALSO_VERDADERO", "", "El ancho de banda de un canal analógico se mide en Hertz (Hz).", "Verdadero", "Falso", "", "", "", "A", "1", 5, "OK"),
        ("RESPUESTA_COMPUESTA", "", "I. La conmutación de paquetes no reserva recursos exclusivos.\nII. La conmutación de circuitos establece un canal dedicado.", "A. Si la primera es verdadera", "B. Si la segunda es verdadera", "C. Si ambas son verdaderas", "D. Si ninguna es verdadera", "", "C", "1", 5, "OK"),
        ("SELECCION_SIMPLE", "", "¿Cuál es la topología física más empleada en redes de área local cableadas modernas?", "Bus", "Anillo", "Estrella", "Malla completa", "Token Ring", "C", "1", 5, "OK"),
        ("FALSO_VERDADERO", "", "Una dirección MAC tiene una longitud estándar de 48 bits.", "Verdadero", "Falso", "", "", "", "A", "1", 5, "OK"),
        ("PREGUNTA_CON_CLAVE", "", "Son protocolos de la capa de transporte en TCP/IP: 1. TCP, 2. UDP, 3. SCTP, 4. ICMP.", "1, 2 y 3 son correctas", "1 y 3 son correctas", "2 y 4 son correctas", "Solo 4 es correcta", "Todas son correctas", "A", "1", 5, "OK"),
        ("SELECCION_SIMPLE", "", "¿Qué dispositivo conmuta tramas basándose en las direcciones MAC de origen y destino?", "Hub", "Switch", "Repetidor", "Transceptor", "Atenuador", "B", "1", 5, "OK")
    ]

    fill_facil = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_medio = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    fill_dificil = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for r in faciles_data:
        ws_banco.append(list(r))
        curr_row = ws_banco.max_row
        for col_i in range(1, len(headers) + 1):
            cell = ws_banco.cell(row=curr_row, column=col_i)
            cell.border = thin_border
            cell.alignment = align_top_wrap
            if col_i == 10:
                cell.fill = fill_facil

    # 30 Medias (Dificultad 2)
    for i in range(1, 31):
        tipo_op = ["SELECCION_SIMPLE", "RESPUESTA_COMPUESTA", "PREGUNTA_CON_CLAVE", "FALSO_VERDADERO"][(i-1) % 4]
        if tipo_op == "FALSO_VERDADERO":
            row = [tipo_op, "", f"Pregunta Media {i}: En la modulación QPSK cada símbolo transmitido representa exactamente 2 bits.", "Verdadero", "Falso", "", "", "", "A", "2", 5, "OK"]
        elif tipo_op == "RESPUESTA_COMPUESTA":
            row = [tipo_op, "", f"Pregunta Media {i}: I. El retardo de propagación depende de la distancia y velocidad de la luz.\nII. El retardo de transmisión depende de la tasa de bits.", "A. Si la primera es verdadera", "B. Si la segunda es verdadera", "C. Si ambas son verdaderas", "D. Si ninguna es verdadera", "", "C", "2", 5, "OK"]
        elif tipo_op == "PREGUNTA_CON_CLAVE":
            row = [tipo_op, "", f"Pregunta Media {i}: Ventajas de la multiplexación OFDM: 1. Alta eficiencia espectral, 2. Robustez ante multitrayecto, 3. Baja ISI, 4. Ausencia total de distorsión.", "1, 2 y 3 son correctas", "1 y 3 son correctas", "2 y 4 son correctas", "Solo 4 es correcta", "Todas son correctas", "A", "2", 5, "OK"]
        else:
            row = [tipo_op, "", f"Pregunta Media {i}: ¿Qué protocolo de enrutamiento dinámico utiliza el algoritmo de vector de distancias Bellman-Ford?", "OSPF", "RIP", "BGP", "IS-IS", "EIGRP", "B", "2", 5, "OK"]
        
        ws_banco.append(row)
        curr_row = ws_banco.max_row
        for col_i in range(1, len(headers) + 1):
            cell = ws_banco.cell(row=curr_row, column=col_i)
            cell.border = thin_border
            cell.alignment = align_top_wrap
            if col_i == 10:
                cell.fill = fill_medio

    # 15 Difíciles (Dificultad 3)
    for i in range(1, 16):
        if i <= 5:
            row = ["PROBLEMA", f"CASO-0{i}", f"Problema Difícil {i}: Para un radioenlace a 5 GHz con distancia d=10 km, determine la atenuación FSPL: $ FSPL = 20 log(d) + 20 log(f) + 92.45 $", "112.4 dB", "126.4 dB", "140.2 dB", "98.5 dB", "150.0 dB", "B", "3", 5, "OK"]
        elif i <= 10:
            row = ["SUBPROBLEMA", f"CASO-0{i-5}", f"Subproblema {i}: Con la atenuación FSPL anterior, calcule la potencia recibida en dBm si la potencia Tx es 20 dBm y ganancias de antena 20 dBi cada una:", "-66.4 dBm", "-76.4 dBm", "-86.4 dBm", "-96.4 dBm", "-56.4 dBm", "A", "3", 5, "OK"]
        else:
            row = ["SELECCION_SIMPLE", "", f"Pregunta Difícil {i}: En una modulación 256-QAM con ancho de banda de 20 MHz y factor roll-off 0.25, la tasa binaria neta teórica es:", "128 Mbps", "106.6 Mbps", "160 Mbps", "80 Mbps", "64 Mbps", "A", "3", 5, "OK"]
        
        ws_banco.append(row)
        curr_row = ws_banco.max_row
        for col_i in range(1, len(headers) + 1):
            cell = ws_banco.cell(row=curr_row, column=col_i)
            cell.border = thin_border
            cell.alignment = align_top_wrap
            if col_i == 10:
                cell.fill = fill_dificil

    # Sheet 3: Ejemplos
    ws_ej = wb_valido.create_sheet(title="Ejemplos")
    ws_ej.append(headers)
    for col_num in range(1, len(headers) + 1):
        c = ws_ej.cell(row=1, column=col_num)
        c.font = header_font
        c.fill = header_fill_cyan
        c.alignment = align_center

    for idx, width in enumerate(col_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(idx)
        ws_ej.column_dimensions[col_letter].width = width

    ws_ej.append(["FALSO_VERDADERO", "", "El agua hierve a 100 grados Celsius al nivel del mar.", "Verdadero", "Falso", "", "", "", "A", "1", 5, "OK"])
    ws_ej.append(["SELECCION_SIMPLE", "", "¿Qué órgano bombea la sangre en el cuerpo humano?", "Pulmón", "Hígado", "Corazón", "Estómago", "Riñón", "C", "2", 5, "OK"])
    ws_ej.append(["RESPUESTA_COMPUESTA", "", "I. La capital de Bolivia es Sucre.\nII. La sede de gobierno es La Paz.", "A. Si la primera es verdadera", "B. Si la segunda es verdadera", "C. Si ambas son verdaderas", "D. Si ninguna es verdadera", "", "C", "2", 5, "OK"])

    for d in output_dirs:
        filepath = os.path.join(d, "BANCO_PRUEBA_VALIDO_60PREGUNTAS.xlsx")
        wb_valido.save(filepath)
        print(f"[OK] Archivo Valido generado en: {filepath}")

    # =========================================================================
    # 2. EXCEL INVÁLIDO (25 PREGUNTAS CON ERRORES DE CUOTA Y FORMATO)
    # =========================================================================
    wb_invalido = openpyxl.Workbook()
    
    ws_inst_inv = wb_invalido.active
    ws_inst_inv.title = "Instrucciones"
    ws_inst_inv.append(["BANCO CON ERRORES PARA PRUEBAS UNITARIAS"])

    ws_banco_inv = wb_invalido.create_sheet(title="Banco")
    ws_banco_inv.append(headers)
    for col_num in range(1, len(headers) + 1):
        c = ws_banco_inv.cell(row=1, column=col_num)
        c.font = header_font
        c.fill = header_fill_red
        c.alignment = align_center

    for idx, width in enumerate(col_widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(idx)
        ws_banco_inv.column_dimensions[col_letter].width = width

    # Errores intencionales
    ws_banco_inv.append(["FALSO_VERDADERO", "", "Pregunta Falso/Verdadero con clave erronea (C)", "Verdadero", "Falso", "", "", "", "C", "1", 5, "Error: Clave debe ser A o B"])
    ws_banco_inv.append(["SELECCION_SIMPLE", "", "", "Distractor A", "Distractor B", "Distractor C", "Distractor D", "Distractor E", "A", "1", 5, "Error: Falta enunciado"])
    ws_banco_inv.append(["SELECCION_SIMPLE", "", "Pregunta de seleccion con solo 2 opciones", "Opcion 1", "Opcion 2", "", "", "", "A", "2", 5, "Error: Requiere al menos 4 opciones"])
    ws_banco_inv.append(["SUBPROBLEMA", "", "Subproblema sin grupo de caso padre asignado", "Distractor A", "Distractor B", "Distractor C", "Distractor D", "Distractor E", "B", "3", 5, "Error: Falta grupo"])
    ws_banco_inv.append(["RESPUESTA_COMPUESTA", "", "I. Premisa 1.\nII. Premisa 2.", "A", "B", "C", "D", "", "Z", "2", 5, "Error: Clave debe ser A, B, C o D"])

    for i in range(6, 26):
        ws_banco_inv.append(["SELECCION_SIMPLE", "", f"Pregunta incompleta {i}", "Opcion A", "Opcion B", "Opcion C", "Opcion D", "Opcion E", "A", "2", 5, "OK"])

    for d in output_dirs:
        filepath = os.path.join(d, "BANCO_PRUEBA_CON_ERRORES.xlsx")
        wb_invalido.save(filepath)
        print(f"[ERROR] Archivo Invalido generado en: {filepath}")

if __name__ == "__main__":
    generate_sample_excels()
